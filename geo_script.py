import os
import openpyxl
import googlemaps

key1 = "AIzaSyCoNxGcGm-rX-6MnY0T6DPvGlOR8D-nNpM"
key2 = "AIzaSyCn1oqUGAj6oa7A4e86CNWgqCdlIdCYbgg"
key3 = "AIzaSyCLWxnRgMEXoNlDP7wzuPFj3SF0PmOsHfo"
key4 = "AIzaSyBpsRVqjjgknBOvQRnwWQ5u2Jj0JR6Dzms"
key5 = "AIzaSyDazpzJ5cdA7W75AWBGZ7pD0QrXBzwhUBA"
key6 = "AIzaSyCGaa_fPnEAIc0hhb6tyAiQsUCHdfaAzqA"
key7 = "AIzaSyBGRqSKCxYx1pceNCge8LevV8q_gS9D4a0"


print("When the script is done running this window will close :)")


#If you are getting an error try changing the key out with a different key from
# the choices above (i.e key1, key2, key3, key4, key5)
gm = googlemaps.Client(key = key7)
wb = openpyxl.load_workbook(filename = 'address.xlsx')

sheet = wb.active

all_addy = []
master = []

for i in sheet.iter_rows():
    if i[0].value == 'Address':
        continue
    
    word = i[0].value + " " + i[1].value

    geocode_result = gm.geocode(word)[0]

    lat = geocode_result["geometry"]["location"]["lat"]
    lng = geocode_result["geometry"]["location"]["lng"]
    master.append((word, lat, lng))


wb = openpyxl.load_workbook(filename = 'address.xlsx')

she = wb.create_sheet(title = "lat-long")

i = 0

for row in she.iter_rows(min_row = 1, max_col = 3, max_row = len(master)):
    y = 0
    for cell in row:
        cell.value = master[i][y]
        y = y + 1
        
    i = i + 1
    
wb.save("address.xlsx")




    
